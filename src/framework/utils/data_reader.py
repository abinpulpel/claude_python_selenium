"""Generic test-data readers for JSON, YAML, CSV, and Excel sources.

Each method returns a list of dictionaries (one per record) so callers can
feed results directly into ``pytest.mark.parametrize``.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from framework.exceptions.framework_exceptions import FrameworkException


class DataReader:
    """Reads structured test data from common file formats."""

    @staticmethod
    def read_json(path: str) -> list[dict[str, Any]]:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        return data if isinstance(data, list) else [data]

    @staticmethod
    def read_yaml(path: str) -> list[dict[str, Any]]:
        data = yaml.safe_load(Path(path).read_text(encoding="utf-8"))
        return data if isinstance(data, list) else [data]

    @staticmethod
    def read_csv(path: str) -> list[dict[str, Any]]:
        with open(path, newline="", encoding="utf-8") as handle:
            return list(csv.DictReader(handle))

    @staticmethod
    def read_excel(path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers, *records = rows
        return [dict(zip(headers, record)) for record in records]

    @classmethod
    def read(cls, path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
        """Dispatch to the correct reader based on file extension."""
        suffix = Path(path).suffix.lower()
        readers = {
            ".json": lambda: cls.read_json(path),
            ".yaml": lambda: cls.read_yaml(path),
            ".yml": lambda: cls.read_yaml(path),
            ".csv": lambda: cls.read_csv(path),
            ".xlsx": lambda: cls.read_excel(path, sheet_name),
        }
        reader = readers.get(suffix)
        if reader is None:
            raise FrameworkException(f"Unsupported data file extension: {suffix}")
        return reader()
